"""
Génère documentation_data_securite.xlsx
Script autonome — exécuter depuis la racine du projet.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _risk_fill(level: str) -> PatternFill:
    colors = {
        "CRITIQUE": "FF4444",
        "ÉLEVÉ":    "FF8C00",
        "MOYEN":    "FFD700",
        "✅":       "00AA44",
        "⚠️":       "FF8C00",
    }
    for key, color in colors.items():
        if key in level:
            return PatternFill("solid", fgColor=color)
    return PatternFill("solid", fgColor="E8E8E8")

def _risk_font(level: str) -> Font:
    if "CRITIQUE" in level or "ÉLEVÉ" in level:
        return Font(bold=True, color="FFFFFF", size=10)
    if "MOYEN" in level:
        return Font(bold=True, color="333333", size=10)
    return Font(size=10, color="333333")

def apply_header(ws, row, cols, values, bg="1F3864", fg="FFFFFF"):
    for col, val in zip(cols, values):
        c = ws.cell(row=row, column=col, value=val)
        c.font = Font(bold=True, color=fg, size=11)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()

def apply_row(ws, row, cols, values, risk_col=None):
    for i, (col, val) in enumerate(zip(cols, values)):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = _border()
        c.font = Font(size=10)
        # Alternance de couleur de fond par ligne
        if row % 2 == 0:
            c.fill = PatternFill("solid", fgColor="F5F8FF")
    # Colonne risque colorée
    if risk_col:
        c = ws.cell(row=row, column=risk_col)
        val = c.value or ""
        c.fill = _risk_fill(val)
        c.font = _risk_font(val)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def add_section_title(ws, row, ncols, title, bg="2E5090"):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 1 — Vue d'ensemble
# ─────────────────────────────────────────────────────────────────────────────

ws1 = wb.active
ws1.title = "Vue d'ensemble"

HEADERS = ["Script / Fichier", "Rôle", "Menaces prises en charge",
           "Fichiers de suivi", "Classe de risque"]
COL_WIDTHS = [40, 45, 55, 38, 20]

for col, width in enumerate(COL_WIDTHS, 1):
    ws1.column_dimensions[get_column_letter(col)].width = width
ws1.row_dimensions[1].height = 30
ws1.row_dimensions[2].height = 35

# Titre principal
ws1.merge_cells("A1:E1")
c = ws1["A1"]
c.value = "ALFRED — Documentation Sécurité · Script · Rôle · Menace · Fichiers de suivi · Classe de risque"
c.font = Font(bold=True, color="FFFFFF", size=14)
c.fill = PatternFill("solid", fgColor="0D1B2A")
c.alignment = Alignment(horizontal="center", vertical="center")

apply_header(ws1, 2, range(1, 6), HEADERS)

SECTIONS = [
    # (titre section, bg, [(script, rôle, menaces, fichiers, risque), ...])
    ("PIPELINE D'ENTRÉE", "C0392B", [
        ("src/security/input_validator.py",
         "Sanitisation de toutes les entrées utilisateur (normalisation unicode NFKC, null bytes, 25 patterns compilés)",
         "SQL injection · XSS · Path traversal · Command injection · Prompt injection · SSRF · Homoglyphes · Null bytes",
         "logs/security/security.log",
         "CRITIQUE — OWASP A03"),
        ("src/security/threat_detector.py",
         "Score de menace sur chaque input (keywords + anomalies comportementales)",
         "SQL tautologies · XSS · Command shells · Prompt jailbreak · LDAP injection · Null bytes · Encodage suspect",
         "logs/security/audit_trail.jsonl",
         "CRITIQUE — OWASP A03 · CWE-20"),
        ("src/security/output_filter.py",
         "Masquage des données sensibles dans les réponses générées",
         "Fuite clés API · JWT · Tokens GitHub · Base64 sensible · Variables d'environnement",
         "logs/security/security.log",
         "ÉLEVÉ — OWASP A02"),
    ]),
    ("AUTHENTIFICATION ET CONTRÔLE D'ACCÈS", "8E44AD", [
        ("src/auth/authenticator.py",
         "Auth PIN locale · bcrypt rounds=12 · sel aléatoire 16 bytes · rate limiter intégré",
         "Brute-force · Rainbow tables · Attaque dictionnaire · Réutilisation de hash",
         "data/security/pins.json · logs/security/security.log",
         "CRITIQUE — OWASP A07 · CWE-307"),
        ("src/security/rate_limiter.py",
         "Limitation de taux avec backoff exponentiel (30s→60s→120s) · fenêtre glissante 60s · thread-safe",
         "Brute-force · Credential stuffing · Déni de service local",
         "logs/security/security.log",
         "ÉLEVÉ — OWASP A04 · CWE-307"),
        ("src/security/session_manager.py",
         "Gestion cycle de vie des sessions (création, validité, fermeture, blocage)",
         "Session hijacking · Session fixation · Réutilisation après expiration",
         "Mémoire (migration SQLite prévue)",
         "ÉLEVÉ — OWASP A07 · CWE-613"),
        ("src/security/mfa_manager.py",
         "Authentification multi-facteurs pondérée (face=3 · voix=2 · appareil=2 · PIN=1)",
         "Usurpation d'identité · Compromission d'un seul facteur",
         "config/security/security_settings.json",
         "ÉLEVÉ — OWASP A07"),
        ("src/security/access_control.py",
         "Vérification des permissions par rôle (RBAC)",
         "Escalade de privilèges · Accès non autorisé à des ressources",
         "logs/security/security.log",
         "CRITIQUE — OWASP A01"),
        ("src/security/permission_manager.py",
         "Matrice de permissions — 7 rôles de OWNER à EMERGENCY",
         "Accès latéral · Privilege creep",
         "config/security/roles_permissions.json",
         "CRITIQUE — OWASP A01"),
    ]),
    ("POLITIQUE ZERO TRUST", "1A5276", [
        ("src/security/zero_trust_orchestrator.py",
         "Orchestration 6 étapes : input → menace → appareil → accès → politique → audit · fail-secure",
         "Toutes menaces combinées — aucune étape ne peut être contournée isolément",
         "audit_trail.jsonl · security.log",
         "CRITIQUE — Architecture ZT"),
        ("src/security/policy_engine.py",
         "Évaluation des règles (sensibilité ressource, action, rôle)",
         "Accès ressources CRITICAL par rôles non autorisés · Actions destructives",
         "config/security/zero_trust_rules.json",
         "CRITIQUE — OWASP A01"),
        ("src/security/policy_enforcement_point.py",
         "Application de la décision PDP · Décisions inconnues → refus systématique (fail-secure)",
         "Décisions ambiguës · Tentatives de contournement",
         "logs/security/audit_trail.jsonl",
         "ÉLEVÉ — CWE-636"),
        ("src/security/device_registry.py",
         "Registre des appareils de confiance · Vérification · Révocation · last_seen",
         "Appareil non reconnu · Appareil compromis · Connexion depuis matériel inconnu",
         "data/security/trusted_devices.json · security.log",
         "ÉLEVÉ — OWASP A07 · CWE-287"),
    ]),
    ("CHIFFREMENT ET PROTECTION DES DONNÉES", "117A65", [
        ("src/security/encryption_service.py",
         "Chiffrement Fernet (AES-128-CBC + HMAC-SHA256) · IV aléatoire · Rotation de clé",
         "Lecture données au repos · Exfiltration de fichiers · Falsification de tokens",
         "Clé dans var. d'env. FERNET_KEY (.env exclu Git)",
         "CRITIQUE — OWASP A02 · CWE-311"),
        ("src/security/data_protection.py",
         "Chiffrement sélectif des champs sensibles dans les JSON (password, token, api_key…)",
         "Exposition de secrets dans fichiers de données utilisateur",
         "logs/security/security.log",
         "ÉLEVÉ — OWASP A02 · CWE-312"),
        ("src/security/secret_manager.py",
         "Validation des secrets critiques au démarrage (SECRET_KEY, FERNET_KEY, PIN_SALT)",
         "Démarrage avec secrets par défaut · Secrets manquants",
         ".env (exclu Git via .gitignore)",
         "CRITIQUE — OWASP A05"),
        ("src/security/backup_security.py",
         "Vérification de l'intégrité des sauvegardes",
         "Corruption de backup · Substitution de données · Restauration malveillante",
         "À configurer",
         "MOYEN — CWE-354"),
    ]),
    ("DÉTECTION ET RÉPONSE AUX INCIDENTS", "784212", [
        ("src/security/incident_manager.py",
         "Enregistrement structuré des incidents (level, status, source, timestamp)",
         "Traçabilité événements critiques · Gestion incidents ouverts",
         "data/security/incident_register.json",
         "ÉLEVÉ — OWASP A09"),
        ("src/security/behavioral_detector.py",
         "Détection d'anomalies comportementales (écart vs baseline + score)",
         "Comportement anormal · Dérive de session · Usage atypique",
         "Score en mémoire",
         "MOYEN — CWE-754"),
        ("src/security/prompt_guard.py",
         "Protection contre les injections dans les chaînes LLM",
         "Jailbreak · Extraction de contexte système · Manipulation du LLM",
         "logs/security/security.log",
         "ÉLEVÉ — Spécifique LLM"),
        ("src/security/quarantine_service.py",
         "Isolation des données suspectes",
         "Contamination mémoire · Données corrompues ou malveillantes",
         "À configurer",
         "MOYEN — Architecture défensive"),
    ]),
    ("JOURNALISATION ET TRAÇABILITÉ", "2C3E50", [
        ("src/security/audit_trail.py",
         "Journal JSONL horodaté UTC · user · action · resource · decision",
         "Non-répudiation · Investigation post-incident · Conformité GDPR",
         "logs/security/audit_trail.jsonl",
         "ÉLEVÉ — OWASP A09 · GDPR Art.30"),
        ("src/security/security_logger.py",
         "Logger dédié (INFO/WARNING/ERROR/CRITICAL) · log_access() · log_auth()",
         "Traçabilité auth et accès",
         "logs/security/security.log",
         "ÉLEVÉ — OWASP A09"),
        ("src/security/compliance_manager.py",
         "Suivi de la conformité réglementaire",
         "Non-conformité GDPR · Drift de configuration",
         "config/security/audit_retention_policy.json",
         "MOYEN — GDPR · ISO 27001"),
    ]),
    ("GOUVERNANCE ET MONITORING", "1B4F72", [
        ("src/security/security_dashboard.py",
         "Score sécurité 0-100 (grade A→F) · Conformité GDPR/OWASP · Résumé menaces/incidents · Rapport",
         "Posture dégradée non détectée · Accumulation silencieuse de risques",
         "audit_trail.jsonl · incident_register.json",
         "Monitoring transverse"),
        ("src/security/security_governance.py",
         "13 contrôles durcissement · Matrice risques (likelihood × impact) · Reco CRITICAL→LOW",
         "Dérives de configuration · Lacunes non traitées · Risques non prioritisés",
         "Tous fichiers config sécurité",
         "Gouvernance transverse"),
    ]),
]

current_row = 3
for section_title, section_bg, rows in SECTIONS:
    add_section_title(ws1, current_row, 5, f"  {section_title}", bg=section_bg)
    current_row += 1
    ws1.row_dimensions[current_row - 1].height = 20
    for row_data in rows:
        ws1.row_dimensions[current_row].height = 60
        apply_row(ws1, current_row, range(1, 6), row_data, risk_col=5)
        current_row += 1

ws1.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 2 — Tests de pénétration
# ─────────────────────────────────────────────────────────────────────────────

ws2 = wb.create_sheet("Tests de pénétration")
for col, width in enumerate([38, 45, 55, 14, 22], 1):
    ws2.column_dimensions[get_column_letter(col)].width = width

ws2.merge_cells("A1:E1")
c = ws2["A1"]
c.value = "ALFRED — Suite de tests d'intrusion (136 tests · 100% pass)"
c.font = Font(bold=True, color="FFFFFF", size=14)
c.fill = PatternFill("solid", fgColor="0D1B2A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30
ws2.row_dimensions[2].height = 35

apply_header(ws2, 2, range(1, 6), ["Fichier test", "Rôle", "Vecteurs testés", "Résultat", "Classe de risque"])

PENTEST_ROWS = [
    ("tests/security/test_pentest_input.py",
     "Tests d'intrusion — vecteurs d'injection",
     "SQL (8 payloads) · XSS (8) · Path traversal (5) · Prompt injection (8) · Command injection (6) · Edge cases (8) = 43 tests",
     "43/43 ✅",
     "CRITIQUE — OWASP A03"),
    ("tests/security/test_pentest_auth.py",
     "Tests d'intrusion — authentification et rate limiting",
     "Brute-force · Backoff exponentiel · Lockout · Isolation utilisateurs · Session lifecycle · PIN bcrypt = 20 tests",
     "20/20 ✅",
     "CRITIQUE — OWASP A07"),
    ("tests/security/test_pentest_encryption.py",
     "Tests d'intrusion — chiffrement Fernet et data_protection",
     "Round-trip · IV aléatoire · Falsification token · Protection champs sensibles · Insensibilité casse = 16 tests",
     "16/16 ✅",
     "ÉLEVÉ — OWASP A02"),
    ("tests/security/test_pentest_zero_trust.py",
     "Tests d'intrusion — orchestrateur Zero Trust complet",
     "Payloads malveillants → refus · Appareil inconnu → refus · Permission insuffisante · Ressource critique · Happy path = 13 tests",
     "13/13 ✅",
     "CRITIQUE — Architecture ZT"),
    ("tests/security/test_pentest_dashboard_governance.py",
     "Tests dashboard et gouvernance",
     "Score/grade · Compliance structure · Matrice risques · Idempotence · Filtrage priorité = 44 tests",
     "44/44 ✅",
     "Monitoring"),
]

for i, row_data in enumerate(PENTEST_ROWS, 3):
    ws2.row_dimensions[i].height = 65
    apply_row(ws2, i, range(1, 6), row_data, risk_col=5)

ws2.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 3 — Intégration main.py
# ─────────────────────────────────────────────────────────────────────────────

ws3 = wb.create_sheet("Intégration main.py")
for col, width in enumerate([30, 35, 35, 35, 20], 1):
    ws3.column_dimensions[get_column_letter(col)].width = width

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "ALFRED — Points d'intégration sécurité dans main.py"
c.font = Font(bold=True, color="FFFFFF", size=14)
c.fill = PatternFill("solid", fgColor="0D1B2A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30
ws3.row_dimensions[2].height = 35

apply_header(ws3, 2, range(1, 6), ["Point d'intégration", "Module appelé", "Déclencheur", "Effet sécurité", "Classe de risque"])

INTEGRATION_ROWS = [
    ("main.py — démarrage",
     "dotenv.load_dotenv()",
     "Lancement ALFRED",
     "Chargement FERNET_KEY et secrets depuis .env avant tout import",
     "CRITIQUE — OWASP A05"),
    ("main.py — démarrage",
     "device_registry.init_default_device()",
     "Lancement ALFRED",
     "Enregistrement de l'appareil local dans le registre de confiance",
     "ÉLEVÉ — OWASP A07"),
    ("main.py — démarrage",
     "audit_trail.write_audit_event()",
     "Lancement ALFRED",
     "Événement STARTUP tracé dans audit_trail.jsonl",
     "ÉLEVÉ — OWASP A09"),
    ("main.py — boucle principale",
     "input_validator.sanitize_input()",
     "Chaque saisie utilisateur",
     "Rejet des inputs malveillants (25 patterns, unicode, null bytes)",
     "CRITIQUE — OWASP A03"),
    ("main.py — boucle principale",
     "threat_detector.detect_threat()",
     "Après sanitisation",
     "Blocage si score de menace ≥ 3 + message utilisateur + audit",
     "CRITIQUE — OWASP A03"),
    ("main.py — boucle principale",
     "output_filter.filter_output()",
     "Chaque réponse générée",
     "Masquage clés/tokens avant affichage, TTS et stockage mémoire",
     "ÉLEVÉ — OWASP A02"),
    ("main.py — boucle principale",
     "audit_trail.write_audit_event()",
     "Chaque échange",
     "Trace CONVERSATION/ALLOW dans audit_trail.jsonl",
     "ÉLEVÉ — OWASP A09"),
    ("main.py — commande 'reset'",
     "audit_trail.write_audit_event()",
     "Commande de suppression mémoire",
     "Trace DELETE_DATA/session_memory dans l'audit trail",
     "ÉLEVÉ — OWASP A09"),
    ("main.py — fallback dégradé",
     "_SECURITY_AVAILABLE flag",
     "Si imports sécurité échouent",
     "Sanitisation basique active, ALFRED continue sans planter",
     "MOYEN — Résilience"),
]

for i, row_data in enumerate(INTEGRATION_ROWS, 3):
    ws3.row_dimensions[i].height = 55
    apply_row(ws3, i, range(1, 6), row_data, risk_col=5)

ws3.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 4 — Conformité OWASP
# ─────────────────────────────────────────────────────────────────────────────

ws4 = wb.create_sheet("Conformité OWASP Top 10")
for col, width in enumerate([22, 28, 55, 15], 1):
    ws4.column_dimensions[get_column_letter(col)].width = width

ws4.merge_cells("A1:D1")
c = ws4["A1"]
c.value = "ALFRED — Synthèse conformité OWASP Top 10"
c.font = Font(bold=True, color="FFFFFF", size=14)
c.fill = PatternFill("solid", fgColor="0D1B2A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30
ws4.row_dimensions[2].height = 35

apply_header(ws4, 2, range(1, 5), ["Classe OWASP", "Titre", "Modules couvrants", "Statut"])

OWASP_ROWS = [
    ("A01 — Contrôle d'accès",
     "Broken Access Control",
     "access_control · permission_manager · role_manager · policy_engine · zero_trust_orchestrator · roles_permissions.json",
     "✅ Couvert"),
    ("A02 — Exposition données",
     "Cryptographic Failures",
     "encryption_service · data_protection · output_filter · secret_manager",
     "✅ Couvert"),
    ("A03 — Injection",
     "Injection",
     "input_validator (25 patterns) · threat_detector · prompt_guard",
     "✅ Couvert"),
    ("A04 — Design non sécurisé",
     "Insecure Design",
     "rate_limiter (backoff) · policy_enforcement_point (fail-secure) · zero_trust_orchestrator",
     "✅ Couvert"),
    ("A05 — Mauvaise configuration",
     "Security Misconfiguration",
     "secret_manager · security_governance (13 contrôles) · roles_permissions.json · zero_trust_rules.json",
     "✅ Couvert"),
    ("A06 — Composants vulnérables",
     "Vulnerable Components",
     "pip-audit (prévu en CI) · requirements.txt",
     "⚠️ Partiel"),
    ("A07 — Auth et identification",
     "Auth Failures",
     "authenticator (bcrypt) · session_manager · mfa_manager · device_registry · rate_limiter",
     "✅ Couvert"),
    ("A08 — Intégrité logicielle",
     "Software Integrity Failures",
     "backup_security",
     "⚠️ À compléter"),
    ("A09 — Logging et monitoring",
     "Security Logging Failures",
     "audit_trail · security_logger · incident_manager · security_dashboard",
     "✅ Couvert"),
    ("A10 — SSRF",
     "Server-Side Request Forgery",
     "input_validator (patterns localhost / ldap:// / gopher://)",
     "✅ Couvert"),
]

for i, row_data in enumerate(OWASP_ROWS, 3):
    ws4.row_dimensions[i].height = 50
    apply_row(ws4, i, range(1, 5), row_data, risk_col=4)

ws4.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# FEUILLE 5 — Configuration sécurité
# ─────────────────────────────────────────────────────────────────────────────

ws5 = wb.create_sheet("Configuration sécurité")
for col, width in enumerate([42, 45, 45, 30, 20], 1):
    ws5.column_dimensions[get_column_letter(col)].width = width

ws5.merge_cells("A1:E1")
c = ws5["A1"]
c.value = "ALFRED — Fichiers de configuration sécurité"
c.font = Font(bold=True, color="FFFFFF", size=14)
c.fill = PatternFill("solid", fgColor="0D1B2A")
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 30
ws5.row_dimensions[2].height = 35

apply_header(ws5, 2, range(1, 6), ["Fichier", "Contenu", "Menaces prises en charge", "Utilisé par", "Classe de risque"])

CONFIG_ROWS = [
    ("config/security/roles_permissions.json",
     "Matrice RBAC — 7 rôles · permissions · MFA requis · timeout session",
     "Escalade de privilèges · Accès non autorisé",
     "permission_manager.py · access_control.py",
     "CRITIQUE — OWASP A01"),
    ("config/security/zero_trust_rules.json",
     "8 règles Zero Trust (ZT-001→ZT-007 + DEFAULT)",
     "Accès ressources critiques · Actions destructives non autorisées",
     "policy_engine.py",
     "CRITIQUE — Architecture ZT"),
    ("config/security/security_settings.json",
     "Méthode de chiffrement · MFA requis (true)",
     "MFA contourné · Chiffrement faible",
     "mfa_manager.py",
     "ÉLEVÉ — OWASP A05"),
    ("config/security/audit_retention_policy.json",
     "Politique de rétention des logs d'audit",
     "Suppression non autorisée de journaux · Non-conformité GDPR",
     "compliance_manager.py",
     "MOYEN — GDPR Art.5"),
    ("config/security/trusted_devices.json",
     "Configuration initiale du registre d'appareils",
     "Appareils non autorisés",
     "device_registry.py",
     "ÉLEVÉ — OWASP A07"),
    (".env  (exclu de Git)",
     "FERNET_KEY · SECRET_KEY · PIN_SALT · SESSION_TIMEOUT · MAX_LOGIN_ATTEMPTS",
     "Exposition de secrets · Démarrage non sécurisé",
     "encryption_service.py · secret_manager.py · authenticator.py",
     "CRITIQUE — OWASP A05"),
    (".gitignore",
     "Exclusions : .env · *.key · *.pem · data/security/pins.json",
     "Commit accidentel de secrets · Exposition dans l'historique Git",
     "Git",
     "CRITIQUE — OWASP A05"),
]

for i, row_data in enumerate(CONFIG_ROWS, 3):
    ws5.row_dimensions[i].height = 55
    apply_row(ws5, i, range(1, 6), row_data, risk_col=5)

ws5.freeze_panes = "A3"


# ─────────────────────────────────────────────────────────────────────────────
# Sauvegarde
# ─────────────────────────────────────────────────────────────────────────────

OUTPUT = "docs/documentation_data_securite.xlsx"
wb.save(OUTPUT)
print(f"Fichier généré : {OUTPUT}")
